import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta, date
import plotly.graph_objects as go
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="ABD Resource Gap Tool",
    page_icon="🏗️",
    layout="wide"
)

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

HOURS_PER_PERSON_DAY = 10

DISCIPLINE_MAP = {
    "KF-02": {"label": "Civil",       "roles": ["Civils", "Civil Supervisor"]},
    "KF-03": {"label": "Structural",  "roles": ["Boilermaker", "Boilermaker - LH", "Boilermaker  - LH", "Rigger", "Trade Assistant", "Trade Assistant - LH"]},
    "KF-04": {"label": "Mechanical",  "roles": ["Fitter - Mechanical", "Fitter - Mechanical - LH", "Fitter - Mechanical  - LH", "Fitter - Machanical", "Fitter - Mechnical", "Trade Assistant", "Trade Assistant - LH"]},
    "KF-05": {"label": "Piping",      "roles": ["Pipe Fitter", "PolyWelder", "Operator/Poly Welder", "polywelder", "Trade Assistant", "Trade Assistant - LH"]},
    "KF-07": {"label": "Instruments", "roles": ["Technician", "Inspector"]},
    "KF-08": {"label": "Scaffolding", "roles": ["Scaffolder", "Scaffolder TA", "Scaffold Supervisor"]},
}

SUBCONTRACTED = {
    "KF-01": "Earthworks (CAT — subcontractor)",
    "KF-06": "Electrical (subcontractor)",
}

ON_SITE_STATUSES = {"SITE", "SITE-NS"}

SKIP_ROLES = {
    "Category","SITE","SITE-NS","RR","WFH","LEAVE","RDO","TOIL","TRAINING",
    "TRAINING - OD","FI/AM","FO/PM","DI/AM","DO/PM","FI/PM","FO/AM",
    "PUBLIC HOL","local","DI/PM","DO/AM","FI/AM FO/PM","FI/AM/VRR",
    "FI/PM/VRR","SL-ONSITE","CORP OFFICE"
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def parse_hours(val):
    if val is None: return 0.0
    s = str(val).strip()
    try:
        if s.endswith("h"): return float(s[:-1])
        if s.endswith("d"): return float(s[:-1]) * 8
        return float(s)
    except: return 0.0

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, str):
        s = val.replace(" A", "").strip()
        for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d"):
            try: return datetime.strptime(s, fmt).date()
            except: continue
    return None

def week_label(dt):
    return datetime.combine(dt, datetime.min.time()).strftime("%G-W%V")

def all_days_between(start, finish):
    """7-day site — every calendar day counts."""
    days = []
    cur = start
    while cur <= finish:
        days.append(cur)
        cur += timedelta(days=1)
    return days

# ─────────────────────────────────────────────
# LOADERS
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_p6(file_bytes):
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=True)
    ws = wb["Full Schedule"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        act_id     = row[0]
        act_name   = row[1]
        status     = row[2]
        start      = parse_date(row[5])
        finish     = parse_date(row[6])
        tot_float  = parse_hours(row[10])
        rem_hours  = parse_hours(row[13])
        discipline = str(row[24]).strip() if row[24] else None

        if not act_id or not discipline: continue
        if discipline not in DISCIPLINE_MAP and discipline not in SUBCONTRACTED: continue
        if status == "Completed": continue
        if start is None or finish is None: continue

        rows.append({
            "activity_id":   str(act_id).strip(),
            "activity_name": str(act_name).strip() if act_name else "",
            "status":        status or "",
            "start":         start,
            "finish":        finish,
            "total_float_h": tot_float,
            "rem_hours":     rem_hours,
            "discipline":    discipline,
        })
    wb.close()
    return pd.DataFrame(rows)

@st.cache_data(show_spinner=False)
def load_roster(file_bytes):
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb["Kaefer"]

    col_dates = {}
    for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
        for j, val in enumerate(row):
            if isinstance(val, datetime):
                col_dates[j] = val.date()

    people = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        name = row[0]
        role = str(row[1]).strip() if row[1] else None
        if not name or not role or role in SKIP_ROLES: continue
        row_list = list(row)
        for j, dt in col_dates.items():
            if j >= len(row_list): continue
            status = str(row_list[j]).strip() if row_list[j] else ""
            if status in ON_SITE_STATUSES:
                people.append({
                    "name": str(name).strip(),
                    "role": role,
                    "date": dt,
                    "week": week_label(dt),
                })
    wb.close()
    return pd.DataFrame(people)

# ─────────────────────────────────────────────
# ANALYSIS
# ─────────────────────────────────────────────

def build_demand(p6_df, analysis_start, analysis_end, hrs_per_day):
    """
    For each task, spread remaining hours evenly across all calendar days
    (7-day site) in the remaining task window. Sum across tasks per day,
    then AVERAGE across days in each week to get avg daily headcount needed.
    Unit = people needed on site simultaneously on an average day that week.
    """
    demand_rows = []
    for _, task in p6_df.iterrows():
        disc = task["discipline"]
        if disc not in DISCIPLINE_MAP: continue
        disc_label = DISCIPLINE_MAP[disc]["label"]
        rem_hours  = task["rem_hours"]
        if rem_hours <= 0: continue

        eff_start = max(task["start"], analysis_start)
        eff_end   = min(task["finish"], analysis_end)
        if eff_start > eff_end: continue

        future_days = all_days_between(eff_start, eff_end)
        if not future_days: continue

        # People needed per day for this task
        daily_people = rem_hours / len(future_days) / hrs_per_day

        for day in future_days:
            demand_rows.append({
                "date":        day,
                "week":        week_label(day),
                "discipline":  disc_label,
                "daily_hc":    daily_people,
            })

    if not demand_rows:
        return pd.DataFrame()

    df = pd.DataFrame(demand_rows)
    # Sum all tasks per day per discipline
    daily = df.groupby(["date", "week", "discipline"])["daily_hc"].sum().reset_index()
    # Average across days in the week → avg daily headcount needed
    demand = daily.groupby(["week", "discipline"])["daily_hc"].mean().reset_index()
    demand.columns = ["week", "discipline", "demand_headcount"]
    demand["demand_headcount"] = demand["demand_headcount"].round(1)
    return demand

def build_supply(roster_df):
    """
    Count unique people on site per day per discipline,
    then average across days in the week → avg daily headcount available.
    Unit = people on site simultaneously on an average day that week.
    """
    if roster_df.empty: return pd.DataFrame()

    role_to_discs = {}
    for info in DISCIPLINE_MAP.values():
        for role in info["roles"]:
            role_to_discs.setdefault(role, set()).add(info["label"])

    expanded = []
    for _, row in roster_df.iterrows():
        for disc in role_to_discs.get(row["role"], set()):
            expanded.append({
                "date":       row["date"],
                "week":       row["week"],
                "discipline": disc,
                "name":       row["name"],
            })

    if not expanded: return pd.DataFrame()

    exp_df = pd.DataFrame(expanded)
    # Unique people per day per discipline
    daily = exp_df.groupby(["date", "week", "discipline"])["name"].nunique().reset_index()
    daily.columns = ["date", "week", "discipline", "daily_hc"]
    # Average across days in the week
    supply = daily.groupby(["week", "discipline"])["daily_hc"].mean().reset_index()
    supply.columns = ["week", "discipline", "supply_headcount"]
    supply["supply_headcount"] = supply["supply_headcount"].round(1)
    return supply

def build_gap(demand_df, supply_df):
    if demand_df.empty: return pd.DataFrame()
    gap = demand_df.merge(supply_df, on=["week", "discipline"], how="left")
    gap["supply_headcount"] = gap["supply_headcount"].fillna(0)
    gap["gap"] = (gap["supply_headcount"] - gap["demand_headcount"]).round(1)
    gap["status"] = gap["gap"].apply(lambda x: "✅ OK" if x >= 0 else "🔴 Short")
    return gap.sort_values(["week", "discipline"])

def build_suggestions(p6_df, gap_df, hrs_per_day):
    if p6_df.empty or gap_df.empty: return []
    shortage_weeks = gap_df[gap_df["gap"] < 0][["week", "discipline", "gap"]].copy()
    if shortage_weeks.empty: return []

    suggestions = []
    seen = set()
    for _, task in p6_df.iterrows():
        disc = task["discipline"]
        if disc not in DISCIPLINE_MAP: continue
        float_h = task["total_float_h"]
        if not float_h or float_h <= 0: continue
        disc_label  = DISCIPLINE_MAP[disc]["label"]
        float_days  = round(float_h / hrs_per_day, 1)
        matches = shortage_weeks[shortage_weeks["discipline"] == disc_label]
        for _, short in matches.iterrows():
            key = task["activity_id"] + short["week"]
            if key in seen: continue
            seen.add(key)
            suggestions.append({
                "Shortage week":   short["week"],
                "Discipline":      disc_label,
                "Gap (people/day)": round(abs(short["gap"]), 1),
                "Activity ID":     task["activity_id"],
                "Activity":        task["activity_name"][:70],
                "Task finishes":   str(task["finish"]),
                "Float avail (d)": float_days,
                "Suggestion":      f"Can defer up to {float_days}d — may ease {short['week']} shortage",
            })
    return suggestions[:60]

# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────

st.title("🏗️ ABD Resource Gap Tool")
st.caption("P6 schedule demand vs Kaefer on-site roster · avg daily headcount · 7-day site")

with st.sidebar:
    st.header("📂 Upload files")
    p6_file     = st.file_uploader("P6 Schedule (.xlsm / .xlsx)", type=["xlsm", "xlsx"])
    roster_file = st.file_uploader("Kaefer Roster (.xlsx)",        type=["xlsx"])

    st.divider()
    st.subheader("⚙️ Settings")
    hrs_per_day    = st.number_input("Hours per person per day", value=10, min_value=6, max_value=12)
    today_override = st.date_input("Analysis start date", value=date.today())
    end_override   = st.date_input("Analysis end date",   value=date(2026, 10, 1))

    st.divider()
    st.subheader("🗂️ Discipline mapping")
    for code, info in DISCIPLINE_MAP.items():
        st.caption(f"**{code}** → {info['label']}")
    st.caption("─")
    for code, label in SUBCONTRACTED.items():
        st.caption(f"**{code}** → ⚠️ {label}")

if not p6_file or not roster_file:
    st.info("👈 Upload both files in the sidebar to get started.")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("**1 · Upload**\nP6 schedule export and Kaefer roster")
    with c2:
        st.markdown("**2 · Analyse**\nMatches discipline → role, compares avg daily headcount demand vs supply by week")
    with c3:
        st.markdown("**3 · Act**\nReview gaps, update roster, re-upload to confirm gaps closed")
    st.stop()

# ─── Load ───
with st.spinner("Reading P6 schedule…"):
    p6_df = load_p6(p6_file.read())

with st.spinner("Reading Kaefer roster…"):
    roster_df = load_roster(roster_file.read())

with st.spinner("Calculating gaps…"):
    demand_df   = build_demand(p6_df, today_override, end_override, hrs_per_day)
    supply_df   = build_supply(roster_df)
    gap_df      = build_gap(demand_df, supply_df)
    suggestions = build_suggestions(p6_df, gap_df, hrs_per_day)

# ─── KPIs ───
total_weeks    = gap_df["week"].nunique()        if not gap_df.empty else 0
shortage_count = int((gap_df["gap"] < 0).sum()) if not gap_df.empty else 0
ok_count       = int((gap_df["gap"] >= 0).sum()) if not gap_df.empty else 0
worst_gap      = round(gap_df["gap"].min(), 1)   if not gap_df.empty else 0

k1, k2, k3, k4 = st.columns(4)
k1.metric("Weeks in analysis",         total_weeks)
k2.metric("Discipline-weeks ✅ OK",    ok_count)
k3.metric("Discipline-weeks 🔴 Short", shortage_count)
k4.metric("Worst gap (people/day)",    worst_gap, delta_color="inverse")

st.caption("ℹ️ Units: **average daily headcount** — people needed on site simultaneously on an average day that week (7-day site)")
st.divider()

# ─── Heatmap ───
st.subheader("Resource gap heatmap")
st.caption("Each cell = avg daily supply minus avg daily demand.  **Green = surplus · Red = shortage**")

if not gap_df.empty:
    pivot = gap_df.pivot_table(
        index="discipline", columns="week", values="gap", aggfunc="sum"
    ).sort_index()

    text_vals = [[f"+{v:.1f}" if v >= 0 else f"{v:.1f}" for v in row] for row in pivot.values]

    fig = go.Figure(go.Heatmap(
        z=pivot.values,
        x=list(pivot.columns),
        y=list(pivot.index),
        text=text_vals,
        texttemplate="%{text}",
        colorscale=[
            [0.0,  "#7f1d1d"],
            [0.35, "#ef4444"],
            [0.5,  "#fef2f2"],
            [0.65, "#bbf7d0"],
            [1.0,  "#14532d"],
        ],
        zmid=0,
        colorbar=dict(title="Gap<br>(ppl/day)", thickness=14),
    ))
    fig.update_layout(
        height=max(280, len(pivot.index) * 64 + 120),
        xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
        yaxis=dict(tickfont=dict(size=12)),
        margin=dict(l=10, r=10, t=10, b=100),
    )
    st.plotly_chart(fig, use_container_width=True)

st.divider()

# ─── Bar chart ───
st.subheader("Avg daily demand vs supply by week")

disc_list = sorted(gap_df["discipline"].unique()) if not gap_df.empty else []
sel_disc  = st.selectbox("Discipline", ["All disciplines"] + disc_list)

plot_df = gap_df.copy()
if sel_disc != "All disciplines":
    plot_df = plot_df[plot_df["discipline"] == sel_disc]

if not plot_df.empty:
    agg = plot_df.groupby("week").agg(
        demand=("demand_headcount", "sum"),
        supply=("supply_headcount", "sum"),
    ).reset_index()

    fig2 = go.Figure()
    fig2.add_bar(x=agg["week"], y=agg["demand"], name="Demand (schedule)", marker_color="#dc2626", opacity=0.85)
    fig2.add_bar(x=agg["week"], y=agg["supply"], name="Supply (roster)",   marker_color="#16a34a", opacity=0.85)
    fig2.update_layout(
        barmode="group", height=360,
        xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
        yaxis=dict(title="Avg daily headcount (people)"),
        legend=dict(orientation="h", y=1.12),
        margin=dict(l=10, r=10, t=30, b=100),
    )
    st.plotly_chart(fig2, use_container_width=True)

st.divider()

# ─── Gap table ───
st.subheader("Gap detail")

col_a, col_b = st.columns(2)
with col_a:
    shorts_only = st.checkbox("Show shortages only", value=True)
with col_b:
    wk_filter = st.multiselect("Filter weeks", options=sorted(gap_df["week"].unique()) if not gap_df.empty else [])

tbl = gap_df.copy()
if shorts_only:
    tbl = tbl[tbl["gap"] < 0]
if wk_filter:
    tbl = tbl[tbl["week"].isin(wk_filter)]

if not tbl.empty:
    st.dataframe(
        tbl[["week","discipline","demand_headcount","supply_headcount","gap","status"]]
          .rename(columns={
              "week":             "Week",
              "discipline":       "Discipline",
              "demand_headcount": "Demand (ppl/day)",
              "supply_headcount": "Supply (ppl/day)",
              "gap":              "Gap (ppl/day)",
              "status":           "Status",
          }),
        use_container_width=True, hide_index=True,
    )
else:
    st.success("No shortages found for these filters. ✅")

st.divider()

# ─── Suggestions ───
st.subheader("📋 Schedule adjustment suggestions")
st.caption("Tasks with positive float that could be deferred to ease shortage weeks. Verify against critical path before acting.")

if suggestions:
    st.dataframe(
        pd.DataFrame(suggestions).sort_values(["Shortage week", "Gap (people/day)"], ascending=[True, False]),
        use_container_width=True, hide_index=True,
    )
else:
    st.info("No float-based suggestions — either no shortages or no tasks have positive float.")

st.divider()

# ─── Excluded ───
st.subheader("ℹ️ Excluded disciplines")
for code, label in SUBCONTRACTED.items():
    st.caption(f"• **{code}** — {label}")

st.divider()

# ─── Raw data ───
with st.expander("🔍 P6 tasks loaded"):
    st.dataframe(
        p6_df[["activity_id","activity_name","status","start","finish","rem_hours","discipline","total_float_h"]]
          .sort_values("start"),
        use_container_width=True, hide_index=True,
    )

with st.expander("🔍 Roster — on-site days by person"):
    if not roster_df.empty:
        st.dataframe(
            roster_df.groupby(["name","role","week"]).size()
              .reset_index(name="days_on_site")
              .sort_values(["week","name"]),
            use_container_width=True, hide_index=True,
        )
    else:
        st.warning("No roster data loaded.")
